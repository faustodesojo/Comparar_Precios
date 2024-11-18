import requests
from bs4 import BeautifulSoup
import openpyxl

# Cargar el archivo de Excel
archivo_excel = r"C:\Users\Windows\Desktop\PYTHON\COMPARAR_PRECIOS.xlsm"  # Asegúrate de usar el prefijo 'r' o dobles barras '\\' para rutas en Windows
wb = openpyxl.load_workbook(archivo_excel)
ws = wb.active

# Iterar sobre las filas para obtener URLs
for fila in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=13, max_col=13):  # Columna M = URL
    celda_url = fila[0]
    if celda_url.value:
        url = celda_url.value
        try:
            # Hacer la solicitud a la URL
            response = requests.get(url)
            response.raise_for_status()  # Asegurarse de que la solicitud no devuelva errores HTTP
            soup = BeautifulSoup(response.content, "html.parser")

            # Extraer el precio del span con la clase "price"
            span_precio = soup.find("span", class_="price")  # Ajusta la clase según la página
            precio = span_precio.text.strip() if span_precio else "No encontrado"

            # Guardar el precio en la columna de PRECIO (Columna L)
            ws.cell(row=celda_url.row, column=12, value=precio)  # Columna L = PRECIO

        except Exception as e:
            print(f"Error al procesar {url}: {e}")
            ws.cell(row=celda_url.row, column=12, value="Error")  # Escribe 'Error' en caso de fallo

# Guardar los cambios en el archivo Excel
wb.save(r"C:\Users\Windows\Desktop\PYTHON\COMPARAR_PRECIOS.xlsm" )
print("Actualización completada. Archivo guardado como 'archivo_actualizado.xlsx'.")

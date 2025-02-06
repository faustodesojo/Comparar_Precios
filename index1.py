import aiohttp
import asyncio
import ssl
from bs4 import BeautifulSoup
import openpyxl

# Cargar el archivo de Excel
archivo_excel = r"C:\Users\Windows\Desktop\PYTHON\COMPARAR_PRECIOS.xlsx"
wb = openpyxl.load_workbook(archivo_excel)
ws = wb.active

# Configurar SSL seguro para evitar bloqueos
ssl_context = ssl.create_default_context()
ssl_context.check_hostname = False
ssl_context.verify_mode = ssl.CERT_NONE

# Función para procesar una sola URL
async def procesar_url(session, url, fila):
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept-Language": "es-ES,es;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive"
        }

        async with session.get(url, headers=headers, ssl=ssl_context, timeout=10) as response:
            if response.status != 200:
                ws.cell(row=fila, column=12, value=f"Error HTTP {response.status}")
                return

            html = await response.text()
            soup = BeautifulSoup(html, "html.parser")

            # Extraer precio
            span_precio = soup.find("span", class_="price")
            if span_precio:
                precio_texto = (
                    span_precio.text.strip()
                    .replace("$", "")
                    .replace(".", "")
                    .replace(",", ".")
                )
                precio_numerico = float(precio_texto)
                ws.cell(row=fila, column=12, value=precio_numerico)  # Columna L
            else:
                ws.cell(row=fila, column=12, value="Precio no encontrado")

            # Guardar cambios después de cada actualización
            wb.save(archivo_excel)

    except asyncio.TimeoutError:
        ws.cell(row=fila, column=12, value="Timeout")
    except Exception as e:
        ws.cell(row=fila, column=12, value="Error")
        print(f"Error al procesar {url}: {e}")

# Función principal para manejar múltiples solicitudes
async def main():
    urls = [(fila[0].value, fila[0].row) for fila in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=13, max_col=13) if fila[0].value]
    
    async with aiohttp.ClientSession() as session:
        tareas = [procesar_url(session, url, fila) for url, fila in urls]
        await asyncio.gather(*tareas)

# Ejecutar
asyncio.run(main())

print("Actualización completada. Archivo guardado.")

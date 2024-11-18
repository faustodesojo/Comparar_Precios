import requests
from bs4 import BeautifulSoup
import openpyxl

# Cargar el archivo de Excel
archivo_excel = r"C:\Users\Windows\Desktop\PYTHON\COMPARAR_PRECIOS.xlsx"
wb = openpyxl.load_workbook(archivo_excel)
ws = wb.active

# Iterar sobre las filas para obtener URLs
for fila in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=13, max_col=13):  # Columna M = URL
    celda_url = fila[0]
    if celda_url.value:
        url = celda_url.value
        try:
            # Hacer la solicitud a la URL
            response = requests.get(url)
            response.raise_for_status()  # Asegurarse de que la solicitud no devuelva errores HTTP
            soup = BeautifulSoup(response.content, "html.parser")

            # Extraer el precio del span con la clase "price"
            span_precio = soup.find("span", class_="price")  # Ajusta la clase según la página
            if span_precio:
                # Quitar símbolos y ajustar formato de número
                precio_texto = (
                    span_precio.text.strip()
                    .replace("$", "")     # Eliminar símbolo de dólar
                    .replace(".", "")    # Eliminar separadores de miles
                    .replace(",", ".")   # Convertir separador decimal a punto
                )
                precio_numerico = float(precio_texto)  # Convertir a número
            else:
                precio_numerico = None  # Marcar como vacío si no se encuentra el precio

            # Guardar el precio en la columna de PRECIO (Columna L)
            ws.cell(row=celda_url.row, column=12, value=precio_numerico)  # Columna L = PRECIO

        except Exception as e:
            print(f"Error al procesar {url}: {e}")
            ws.cell(row=celda_url.row, column=12, value="Error")  # Escribe 'Error' en caso de fallo

# Guardar los cambios en el archivo Excel
wb.save(r"C:\Users\Windows\Desktop\PYTHON\COMPARAR_PRECIOS.xlsx")
print("Actualización completada. Archivo guardado.")
